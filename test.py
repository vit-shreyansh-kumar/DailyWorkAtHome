from openpyxl import load_workbook
from itertools import islice
import argparse
import sys


FOUR_STARS = '\u2605\u2605\u2605\u2605'
THREE_STARS = '\u2605\u2605\u2605'
TWO_STARS = '\u2605\u2605'
ONE_STARS = '\u2605'
NA = 'NA'

star_value = {ONE_STARS: 1,
              TWO_STARS: 2,
              THREE_STARS: 3,
              FOUR_STARS: 4}


class BColors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[37m'  # light gray
    FAIL = '\033[91m'  # light red
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class PDFScoreCard:
    FILE_ID = 0
    OVERALL_SCORE = 1
    MISSING_CONTENT = 2
    NAV_REQUIRED = 3
    NAV_PRESENT = 4
    NAV_SCORE = 5
    MH_EXTRA = 6
    MH_MISSING = 7
    SH_EXTRA = 8
    SH_MISSING = 9
    GROUPING = 10
    REFLOW_SCORE = 11
    ORDERING = 12
    TXT_GROUPING = 13
    TXT_STYLE = 14
    LST_INDENT = 15
    LST_LEVELS = 16
    FC_EXTRA = 17
    FC_MISSING = 18
    TABLE_SCORE = 19
    TC_EXTRA = 20
    TC_MISSING = 21
    TC_GROUPING = 22
    TCELL_STYLE = 23
    TCELL_GROUPING = 24


anomalies = []


def print_error(msg):  # pragma: no cover
    """ Print error in red"""
    print('{}{}{}'.format(BColors.FAIL, msg, BColors.ENDC))


def print_warning(msg):  # pragma: no cover
    """ Print warning"""
    print('{}{}{}'.format(BColors.WARNING, msg, BColors.ENDC))


def fatal_error(msg, *args, **kwargs):  # pragma: no cover
    """ Print error in red and exit with error code 1"""
    msg = msg.format(*args, **kwargs)
    print('{}ERROR: {}{}'.format(BColors.FAIL, msg, BColors.ENDC))
    sys.exit(1)


def excel_file(xlsx):  # pragma: no cover
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError as e:
        fatal_error('File {} not found', xlsx)
    except Exception as e:
        fatal_error('File {} is not an Excel file', xlsx)
    return xlsx


def get_sheet(f):  # pragma: no cover
    wb = load_workbook(f)
    if 'Sheet' in wb:
        return wb['Sheet']


def check_scores_stars(file_id, row_num, scores, column):
    score_val = [star_value[s] for s in scores]
    score_val.sort()
    if abs(score_val[-1] - score_val[0]) > 1:
        anomalies.append('Row: {} for File_ID: {} has widely ranging experiences ({} - {}) for {}'.format(row_num,
                                                                                                          file_id,
                                                                                                          scores[0],
                                                                                                          scores[-1],
                                                                                                          column))


def check_scores_stars_na(file_id, row_num, scores, column):
    score_set = set(scores)
    if NA in score_set:
        if (len(score_set) > 1):
            anomalies.append('Row: {} for File_ID: {} seems to show tables for some but not for others'.format(row_num,
                                                                                                               file_id))
    else:
        check_scores_stars(file_id, row_num, scores, column)


def spot_anomalies(excel_files):
    print('{}Looking for anomalies{} ...'.format(BColors.HEADER, BColors.ENDC), end='')
    ws = []
    for e in excel_files:
        s = get_sheet(e)
        if s is None:
            print_error('File %s missing a sheet named \'Sheet\'' % e)
            return
        else:
            ws.append(s)

    row_iters = [islice(sheet.rows, 2, None) for sheet in ws]
    for row_num, row_iter in enumerate(zip(*row_iters)):
        file_ids = set(r[PDFScoreCard.FILE_ID].value for r in row_iter)
        if len(file_ids) > 1:
            fatal_error('Row {} rates different files {!r} across the spreadsheets'.format(row_num+3, list(file_ids)))

        file_id = file_ids.pop()

        overall_scores = [r[PDFScoreCard.OVERALL_SCORE].value for r in row_iter]
        check_scores_stars(file_id, row_num+3, overall_scores, 'OVERALL SCORES')

        nav_scores = [r[PDFScoreCard.NAV_SCORE].value for r in row_iter]
        check_scores_stars(file_id, row_num+3, nav_scores, 'NAVIGATION SCORES')

        reflow_scores = [r[PDFScoreCard.REFLOW_SCORE].value for r in row_iter]
        check_scores_stars(file_id, row_num+3, reflow_scores, 'REFLOW SCORES')

        table_scores = [r[PDFScoreCard.TABLE_SCORE].value for r in row_iter]
        check_scores_stars_na(file_id, row_num+3, table_scores, 'TABLE SCORES')

    if anomalies:
        print_warning('ANOMALIES DETECTED')
        for a in anomalies:
            print_warning(a)
        print_warning('Could you give us more information on what warranted these widely varying experiences? (Please do not amend the scores)')
    else:
        print('{} none found.{}'.format(BColors.OKGREEN, BColors.ENDC))


def run():  # pragma: no cover

    parser = argparse.ArgumentParser(description='Validate the evaluation score files')
    parser.add_argument('-e', dest='excel_files', type=excel_file, required=True, action='append', default=[],
                        help='Evaluation score file')
    args = parser.parse_args()

    spot_anomalies(args.excel_files)


if __name__ == '__main__':  # pragma: no cover
    run()
